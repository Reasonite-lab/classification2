from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw" / "external_candidates"
PROCESSED = ROOT / "data" / "processed"
OUT = ROOT / "reports" / "data_quality"
SOURCE = RAW / "1-s2.0-S0012821X22004290-mmc2.xlsx"
FEATURES = [
    "SIO2(WT%)",
    "TIO2(WT%)",
    "AL2O3(WT%)",
    "FEOT(WT%)",
    "CAO(WT%)",
    "MGO(WT%)",
    "MNO(WT%)",
    "K2O(WT%)",
    "NA2O(WT%)",
    "P2O5(WT%)",
    "V(PPM)",
    "CR(PPM)",
    "NI(PPM)",
    "RB(PPM)",
    "SR(PPM)",
    "Y(PPM)",
    "ZR(PPM)",
    "NB(PPM)",
    "LA(PPM)",
    "CE(PPM)",
    "ND(PPM)",
    "TH(PPM)",
]
MAJORS = FEATURES[:10]
ROW_MAP = {
    "SIO2(WT%)": "SiO2",
    "TIO2(WT%)": "TiO2",
    "AL2O3(WT%)": "Al2O3",
    "CAO(WT%)": "CaO",
    "MGO(WT%)": "MgO",
    "MNO(WT%)": "MnO",
    "K2O(WT%)": "K2O",
    "NA2O(WT%)": "Na2O",
    "P2O5(WT%)": "P2O5",
    "V(PPM)": "V",
    "CR(PPM)": "Cr",
    "NI(PPM)": "Ni",
    "RB(PPM)": "Rb",
    "SR(PPM)": "Sr",
    "Y(PPM)": "Y",
    "ZR(PPM)": "Zr",
    "NB(PPM)": "Nb",
    "LA(PPM)": "La",
    "CE(PPM)": "Ce",
    "ND(PPM)": "Nd",
    "TH(PPM)": "Th",
}


def normalize_name(value: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value).upper())


def digest(path: Path) -> str:
    value = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            value.update(block)
    return value.hexdigest()


def date_text(value: object) -> str | None:
    if value is None or pd.isna(value):
        return None
    parsed = pd.to_datetime(value, errors="coerce")
    if not pd.isna(parsed):
        return parsed.date().isoformat()
    return str(value)


def main() -> None:
    PROCESSED.mkdir(parents=True, exist_ok=True)
    OUT.mkdir(parents=True, exist_ok=True)
    worksheet = load_workbook(SOURCE, data_only=True, read_only=True)["Table S1"]
    labels = {
        str(worksheet.cell(row, 1).value).strip(): row
        for row in range(1, worksheet.max_row + 1)
        if worksheet.cell(row, 1).value is not None
    }

    petdb = pd.read_parquet(PROCESSED / "petdb_primary4_v0_1.parquet")
    petdb_names = {
        normalize_name(part)
        for value in petdb["sample_names"].dropna()
        for part in str(value).split(" || ")
        if normalize_name(part)
    }
    petdb_citations = "\n".join(petdb["citations"].dropna().astype(str)).upper()

    rows: list[dict[str, object]] = []
    rejected_replicates = 0
    rejected_bombs = 0
    for column in range(2, 61):
        sample = worksheet.cell(2, column).value
        if not sample:
            continue
        material_flag = worksheet.cell(9, column).value
        if material_flag == "rep.":
            rejected_replicates += 1
            continue
        if material_flag == "Bomb":
            rejected_bombs += 1
            continue

        record: dict[str, object] = {
            "sample_name": str(sample),
            "source_area": "Cumbre Vieja, La Palma, 2021 eruption",
            "material": "bulk rock lava",
            "sampling_date": date_text(worksheet.cell(3, column).value),
            "lava_date": date_text(worksheet.cell(4, column).value),
            "eruption_day": pd.to_numeric(
                worksheet.cell(5, column).value, errors="coerce"
            ),
            "source_dataset_doi": "10.1016/j.epsl.2022.117793",
            "publication_doi": "10.1016/j.epsl.2022.117793",
            "source_license": "CC BY 4.0",
            "provenance_tier": "publisher_primary_publication_linked",
            "mechanism_cohort": "OCEAN_ISLAND_HOTSPOT",
            "actual_label": "OIB",
            "strict_external_candidate": True,
            "loi_reported": pd.to_numeric(
                worksheet.cell(labels["LOI"], column).value, errors="coerce"
            ),
        }
        for feature, label in ROW_MAP.items():
            record[feature] = pd.to_numeric(
                worksheet.cell(labels[label], column).value, errors="coerce"
            )
        reported_fe2o3 = pd.to_numeric(
            worksheet.cell(labels["Fe2O3"], column).value, errors="coerce"
        )
        record["FE2O3T_REPORTED(WT%)"] = reported_fe2o3
        record["FEOT(WT%)"] = reported_fe2o3 * 0.8998
        rows.append(record)

    out = pd.DataFrame(rows)
    out["normalized_sample_name"] = out["sample_name"].map(normalize_name)
    out["petdb_exact_name_collision"] = out["normalized_sample_name"].isin(
        petdb_names
    )
    out["major_complete"] = out[MAJORS].notna().all(axis=1)
    out["major_total_calc_feot"] = out[MAJORS].sum(
        axis=1, min_count=len(MAJORS)
    )
    reported_major_columns = [feature for feature in MAJORS if feature != "FEOT(WT%)"]
    out["major_total_calc_reported_fe2o3"] = (
        out[reported_major_columns].sum(
            axis=1, min_count=len(reported_major_columns)
        )
        + out["FE2O3T_REPORTED(WT%)"]
    )
    out["major_total_plausible"] = out["major_total_calc_feot"].between(95, 105)
    out["flag_silica_40_55"] = out["SIO2(WT%)"].between(40, 55)
    out["observed_feature_n"] = out[FEATURES].notna().sum(axis=1)
    out["model_ready_common22"] = (
        out["major_complete"]
        & out["major_total_plausible"]
        & out["flag_silica_40_55"]
        & out["observed_feature_n"].eq(len(FEATURES))
    )
    out["strict_external_candidate"] &= out["model_ready_common22"]
    out["record_id"] = "elsevier:2022:lapalma:" + out["sample_name"]
    out.to_parquet(PROCESSED / "la_palma_2021_oib_external_v0_1.parquet", index=False)

    audit = {
        "source_file": str(SOURCE.relative_to(ROOT)),
        "bytes": SOURCE.stat().st_size,
        "sha256": digest(SOURCE),
        "publication_doi": "10.1016/j.epsl.2022.117793",
        "license": "CC BY 4.0",
        "workbook_sheet": "Table S1",
        "raw_lava_columns_after_replicate_and_bomb_exclusion": len(out),
        "excluded_replicate_columns": rejected_replicates,
        "excluded_bomb_columns": rejected_bombs,
        "model_ready_common22": int(out["model_ready_common22"].sum()),
        "feature_nonmissing": out[FEATURES].notna().sum().to_dict(),
        "exact_name_collision_n": int(out["petdb_exact_name_collision"].sum()),
        "publication_detected_in_petdb": any(
            key in petdb_citations
            for key in [
                "10.1016/J.EPSL.2022.117793",
                "MANTLE SOURCE CHARACTERISTICS AND MAGMATIC PROCESSES DURING THE 2021 LA PALMA ERUPTION",
            ]
        ),
        "reported_fe2o3_major_total_range": [
            float(out["major_total_calc_reported_fe2o3"].min()),
            float(out["major_total_calc_reported_fe2o3"].max()),
        ],
        "feot_contract_major_total_range": [
            float(out["major_total_calc_feot"].min()),
            float(out["major_total_calc_feot"].max()),
        ],
        "fe_harmonization": (
            "The source reports total iron as Fe2O3. Convert to total Fe as FeO using "
            "FeOT = 0.8998 * Fe2O3T before applying the PetDB feature contract."
        ),
        "selection_policy": (
            "Use primary analyses of bulk-rock lavas only. Exclude ten replicate columns, "
            "four bomb columns, tephra/ash/lapilli, xenopumice and reference materials."
        ),
        "scope_caveat": (
            "The retained samples form a time series from one 2021 eruption and test "
            "transfer to the Canary ocean-island setting, not province-level independence."
        ),
    }
    (OUT / "la_palma_oib_external_audit.json").write_text(
        json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    pd.DataFrame(
        [
            {
                "candidate": "La Palma 2021 Table S1",
                "target": "OIB",
                "natural_lava_rows": len(out),
                "strict_model_ready_rows": int(out["model_ready_common22"].sum()),
                "publication_overlap_petdb": audit["publication_detected_in_petdb"],
                "exact_name_collision_n": audit["exact_name_collision_n"],
                "shared_feature_max": int(out["observed_feature_n"].max()),
                "status": "strict_external_candidate",
            }
        ]
    ).to_csv(OUT / "la_palma_oib_external_audit.csv", index=False)
    print(json.dumps(audit, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
