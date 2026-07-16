from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw" / "external_candidates"
PROCESSED = ROOT / "data" / "processed"
OUT = ROOT / "reports" / "data_quality"

FEATURES = [
    "SIO2(WT%)",
    "TIO2(WT%)",
    "AL2O3(WT%)",
    "FEOT(WT%)",
    "CAO(WT%)",
    "MGO(WT%)",
    "MNO(WT%)",
    "K2O(WT%)",
    "NA2O(WT%)",
    "P2O5(WT%)",
    "V(PPM)",
    "CR(PPM)",
    "NI(PPM)",
    "RB(PPM)",
    "SR(PPM)",
    "Y(PPM)",
    "ZR(PPM)",
    "NB(PPM)",
    "LA(PPM)",
    "CE(PPM)",
    "ND(PPM)",
    "TH(PPM)",
]
MAJORS = FEATURES[:10]

RGR_ROW_MAP = {
    "SIO2(WT%)": "SiO2",
    "TIO2(WT%)": "TiO2",
    "AL2O3(WT%)": "Al2O3",
    "FEOT(WT%)": "FeO*",
    "CAO(WT%)": "CaO",
    "MGO(WT%)": "MgO",
    "MNO(WT%)": "MnO",
    "K2O(WT%)": "K2O",
    "NA2O(WT%)": "Na2O",
    "P2O5(WT%)": "P2O5",
    "V(PPM)": "V",
    "CR(PPM)": "Cr",
    "NI(PPM)": "Ni",
    "RB(PPM)": "Rb",
    "SR(PPM)": "Sr",
    "Y(PPM)": "Y",
    "ZR(PPM)": "Zr",
    "NB(PPM)": "Nb",
    "LA(PPM)": "La",
    "CE(PPM)": "Ce",
    "ND(PPM)": "Nd",
    "TH(PPM)": "Th",
}


def digest(path: Path) -> str:
    value = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            value.update(block)
    return value.hexdigest()


def normalize_name(value: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value).upper())


def read_pangaea(path: Path) -> tuple[pd.DataFrame, str]:
    text = path.read_text(encoding="utf-8")
    end = text.index("*/")
    header = text[: end + 2]
    frame = pd.read_csv(
        path, sep="\t", skiprows=header.count("\n") + 1, low_memory=False
    )
    return frame, header


def first_numeric(frame: pd.DataFrame, prefixes: list[str]) -> pd.Series:
    result = pd.Series(np.nan, index=frame.index, dtype=float)
    for prefix in prefixes:
        column = next((name for name in frame.columns if name.startswith(prefix)), None)
        if column is not None:
            result = result.fillna(pd.to_numeric(frame[column], errors="coerce"))
    return result


def petdb_reference_state() -> tuple[set[str], str]:
    petdb = pd.read_parquet(PROCESSED / "petdb_primary4_v0_1.parquet")
    names = {
        normalize_name(part)
        for value in petdb["sample_names"].dropna()
        for part in str(value).split(" || ")
        if normalize_name(part)
    }
    citations = "\n".join(petdb["citations"].dropna().astype(str)).upper()
    return names, citations


def finalize_frame(frame: pd.DataFrame) -> pd.DataFrame:
    frame["major_complete"] = frame[MAJORS].notna().all(axis=1)
    frame["major_total_calc"] = frame[MAJORS].sum(axis=1, min_count=len(MAJORS))
    frame["major_total_plausible"] = frame["major_total_calc"].between(95, 105)
    frame["flag_silica_40_55"] = frame["SIO2(WT%)"].between(40, 55)
    frame["observed_feature_n"] = frame[FEATURES].notna().sum(axis=1)
    frame["model_ready_common22"] = (
        frame["flag_silica_40_55"]
        & frame["major_complete"]
        & frame["major_total_plausible"]
        & frame["observed_feature_n"].ge(19)
    )
    return frame


def build_pangaea(petdb_names: set[str]) -> tuple[pd.DataFrame, dict[str, object]]:
    path = RAW / "pangaea_922011.tsv"
    raw, header = read_pangaea(path)
    frame = raw.loc[raw["Material"].eq("whole rock") & raw["Area"].notna()].copy()
    out = pd.DataFrame(index=frame.index)
    out["sample_name"] = frame["Sample label (Sample #)"].astype(str)
    out["source_area"] = frame["Area"].astype(str)
    out["material"] = frame["Material"].astype(str)
    out["rock_description"] = frame["Rock"]
    out["latitude"] = pd.to_numeric(frame["Latitude"], errors="coerce")
    out["longitude"] = pd.to_numeric(frame["Longitude"], errors="coerce")
    out["source_dataset_doi"] = "10.1594/PANGAEA.922011"
    out["publication_doi"] = "10.1029/2020GC008946"
    out["source_license"] = "CC BY"
    out["provenance_tier"] = "repository_primary_publication_linked"
    out["mechanism_cohort"] = np.where(
        out["source_area"].isin(["Erromango", "Vulcan Smt."]),
        "ARC_FRONT",
        "BACK_ARC_TRANSITION",
    )
    out["actual_label"] = np.where(
        out["mechanism_cohort"].eq("ARC_FRONT"), "ARC", pd.NA
    )
    out["strict_external_candidate"] = out["mechanism_cohort"].eq("ARC_FRONT")
    out["normalized_sample_name"] = out["sample_name"].map(normalize_name)
    out["petdb_exact_name_collision"] = out["normalized_sample_name"].isin(petdb_names)
    out["petdb_collision_is_short_ambiguous"] = (
        out["petdb_exact_name_collision"]
        & out["normalized_sample_name"].str.fullmatch(r"E\d{1,2}")
    )

    major_prefixes = {
        "SIO2(WT%)": ["SiO2 [%]"],
        "TIO2(WT%)": ["TiO2 [%]"],
        "AL2O3(WT%)": ["Al2O3 [%]"],
        "FEOT(WT%)": ["Fe oxide [%]"],
        "CAO(WT%)": ["CaO [%]"],
        "MGO(WT%)": ["MgO [%]"],
        "MNO(WT%)": ["MnO [%]"],
        "K2O(WT%)": ["K2O [%]"],
        "NA2O(WT%)": ["Na2O [%]"],
        "P2O5(WT%)": ["P2O5 [%]"],
    }
    trace_prefixes = {
        "V(PPM)": ["V [mg/kg] (Inductively", "V [mg/kg] (X-ray"],
        "CR(PPM)": ["Cr [mg/kg] (Inductively", "Cr [mg/kg] (X-ray"],
        "NI(PPM)": ["Ni [mg/kg] (Inductively", "Ni [mg/kg] (X-ray"],
        "RB(PPM)": ["Rb [mg/kg] (Inductively", "Rb [mg/kg] (X-ray"],
        "SR(PPM)": ["Sr [mg/kg] (Inductively", "Sr [mg/kg] (X-ray"],
        "Y(PPM)": ["Y [mg/kg] (Inductively", "Y [mg/kg] (X-ray"],
        "ZR(PPM)": ["Zr [mg/kg] (Inductively", "Zr [mg/kg] (X-ray"],
        "NB(PPM)": ["Nb [mg/kg] (Inductively", "Nb [mg/kg] (X-ray"],
        "LA(PPM)": ["La [mg/kg] (Inductively"],
        "CE(PPM)": ["Ce [mg/kg] (Inductively"],
        "ND(PPM)": ["Nd [mg/kg] (Inductively"],
        "TH(PPM)": ["Th [mg/kg] (Inductively", "Th [mg/kg] (X-ray"],
    }
    for feature, prefixes in {**major_prefixes, **trace_prefixes}.items():
        out[feature] = first_numeric(frame, prefixes)
    out["major_total_reported"] = first_numeric(frame, ["Total [%]"])
    out["loi_reported"] = first_numeric(frame, ["LOI [%]"])
    out["record_id"] = "pangaea:922011:" + out["sample_name"]
    out = finalize_frame(out.reset_index(drop=True))

    audit = {
        "source_file": str(path.relative_to(ROOT)),
        "bytes": path.stat().st_size,
        "sha256": digest(path),
        "dataset_doi": "10.1594/PANGAEA.922011",
        "publication_doi": "10.1029/2020GC008946",
        "license": "CC BY",
        "header_contains_dataset_citation": "Citation:" in header,
        "header_contains_publication_doi": "10.1029/2020GC008946" in header,
        "whole_rock_n": len(out),
        "arc_front_n": int(out["mechanism_cohort"].eq("ARC_FRONT").sum()),
        "back_arc_transition_n": int(
            out["mechanism_cohort"].eq("BACK_ARC_TRANSITION").sum()
        ),
        "exact_name_collision_n": int(out["petdb_exact_name_collision"].sum()),
        "short_ambiguous_collision_n": int(
            out["petdb_collision_is_short_ambiguous"].sum()
        ),
        "collision_samples": out.loc[
            out["petdb_exact_name_collision"], "sample_name"
        ].tolist(),
        "feature_nonmissing": out[FEATURES].notna().sum().to_dict(),
        "major_total_reported_range": [
            float(out["major_total_reported"].min()),
            float(out["major_total_reported"].max()),
        ],
        "loi_range": [float(out["loi_reported"].min()), float(out["loi_reported"].max())],
        "fe_mapping": (
            "PANGAEA 'Fe oxide [%]' is mapped to FEOT; retain this source-field "
            "mapping as a cross-laboratory harmonization caveat."
        ),
    }
    return out, audit


def build_rgr(petdb_names: set[str]) -> tuple[pd.DataFrame, dict[str, object]]:
    path = RAW / "ggge20719-sup-0002-2014GC005649-ts01.xlsx"
    worksheet = load_workbook(path, data_only=True, read_only=True)["Table 1"]
    labels = {
        re.sub(r"\s+", "", str(worksheet.cell(row, 1).value or "")): row
        for row in range(1, worksheet.max_row + 1)
    }
    feature_rows = {
        feature: labels[re.sub(r"\s+", "", label)]
        for feature, label in RGR_ROW_MAP.items()
    }

    rows: list[dict[str, object]] = []
    location: str | None = None
    for column in range(2, worksheet.max_column + 1):
        location_value = worksheet.cell(2, column).value
        if location_value:
            location = str(location_value)
        sample = worksheet.cell(3, column).value
        if not sample or str(sample).startswith("JB-2"):
            continue
        record: dict[str, object] = {
            "sample_name": str(sample),
            "source_area": location,
            "material": "whole rock",
            "latitude": pd.to_numeric(worksheet.cell(5, column).value, errors="coerce"),
            "longitude": pd.to_numeric(worksheet.cell(6, column).value, errors="coerce"),
            "estimated_age_ma": pd.to_numeric(
                worksheet.cell(4, column).value, errors="coerce"
            ),
            "major_total_reported": pd.to_numeric(
                worksheet.cell(18, column).value, errors="coerce"
            ),
            "loi_reported": np.nan,
            "source_dataset_doi": "10.1002/2014GC005649",
            "publication_doi": "10.1002/2014GC005649",
            "source_license": "publisher_supporting_information",
            "provenance_tier": "publisher_primary_publication_linked",
            "mechanism_cohort": "CONTINENTAL_RIFT",
            "actual_label": "CIB",
            "strict_external_candidate": True,
        }
        for feature, row in feature_rows.items():
            record[feature] = pd.to_numeric(
                worksheet.cell(row, column).value, errors="coerce"
            )
        rows.append(record)

    out = pd.DataFrame(rows)
    out["normalized_sample_name"] = out["sample_name"].map(normalize_name)
    out["petdb_exact_name_collision"] = out["normalized_sample_name"].isin(petdb_names)
    out["petdb_collision_is_short_ambiguous"] = False
    out["record_id"] = "wiley:2014GC005649:" + out["sample_name"]
    out = finalize_frame(out)
    out["strict_external_candidate"] &= out["model_ready_common22"]

    audit = {
        "source_file": str(path.relative_to(ROOT)),
        "bytes": path.stat().st_size,
        "sha256": digest(path),
        "publication_doi": "10.1002/2014GC005649",
        "publisher_table": "Table S1",
        "workbook_sheet": "Table 1",
        "natural_whole_rock_n": len(out),
        "strict_silica_40_55_model_ready_n": int(out["strict_external_candidate"].sum()),
        "silica_excluded_samples": out.loc[
            ~out["flag_silica_40_55"], ["sample_name", "SIO2(WT%)"]
        ].to_dict("records"),
        "exact_name_collision_n": int(out["petdb_exact_name_collision"].sum()),
        "feature_nonmissing": out[FEATURES].notna().sum().to_dict(),
        "normalized_major_total_range": [
            float(out["major_total_calc"].min()),
            float(out["major_total_calc"].max()),
        ],
        "analyzed_total_range": [
            float(out["major_total_reported"].min()),
            float(out["major_total_reported"].max()),
        ],
        "major_note": (
            "The workbook reports normalized major elements and a separate analyzed total. "
            "Model features use the normalized values; analyzed total is retained for QC."
        ),
    }
    return out, audit


def main() -> None:
    PROCESSED.mkdir(parents=True, exist_ok=True)
    OUT.mkdir(parents=True, exist_ok=True)
    petdb_names, petdb_citations = petdb_reference_state()
    pangaea, pangaea_audit = build_pangaea(petdb_names)
    rgr, rgr_audit = build_rgr(petdb_names)

    publication_checks = {
        "pangaea_922011_publication_detected_in_petdb": any(
            key in petdb_citations
            for key in ["2020GC008946", "EVOLUTION OF MAGMATISM IN THE NEW HEBRIDES"]
        ),
        "rgr_publication_detected_in_petdb": any(
            key in petdb_citations
            for key in ["2014GC005649", "BASALT VOLATILE FLUCTUATIONS"]
        ),
    }
    pangaea.to_parquet(
        PROCESSED / "pangaea_922011_arc_backarc_external_v0_1.parquet", index=False
    )
    rgr.to_parquet(PROCESSED / "rio_grande_rift_cib_external_v0_1.parquet", index=False)

    summary = pd.DataFrame(
        [
            {
                "candidate": "PANGAEA.922011 ARC front",
                "target": "ARC",
                "natural_rows": int(pangaea["mechanism_cohort"].eq("ARC_FRONT").sum()),
                "strict_model_ready_rows": int(
                    (
                        pangaea["mechanism_cohort"].eq("ARC_FRONT")
                        & pangaea["model_ready_common22"]
                    ).sum()
                ),
                "publication_overlap_petdb": publication_checks[
                    "pangaea_922011_publication_detected_in_petdb"
                ],
                "exact_name_collision_n": int(
                    pangaea.loc[
                        pangaea["mechanism_cohort"].eq("ARC_FRONT"),
                        "petdb_exact_name_collision",
                    ].sum()
                ),
                "shared_feature_max": int(pangaea["observed_feature_n"].max()),
                "status": "strict_external_candidate",
            },
            {
                "candidate": "PANGAEA.922011 Vate Trough",
                "target": "transition",
                "natural_rows": int(
                    pangaea["mechanism_cohort"].eq("BACK_ARC_TRANSITION").sum()
                ),
                "strict_model_ready_rows": int(
                    (
                        pangaea["mechanism_cohort"].eq("BACK_ARC_TRANSITION")
                        & pangaea["model_ready_common22"]
                    ).sum()
                ),
                "publication_overlap_petdb": publication_checks[
                    "pangaea_922011_publication_detected_in_petdb"
                ],
                "exact_name_collision_n": int(
                    pangaea.loc[
                        pangaea["mechanism_cohort"].eq("BACK_ARC_TRANSITION"),
                        "petdb_exact_name_collision",
                    ].sum()
                ),
                "shared_feature_max": int(pangaea["observed_feature_n"].max()),
                "status": "mechanism_pressure_test",
            },
            {
                "candidate": "Rio Grande Rift Table S1",
                "target": "CIB",
                "natural_rows": len(rgr),
                "strict_model_ready_rows": int(rgr["strict_external_candidate"].sum()),
                "publication_overlap_petdb": publication_checks[
                    "rgr_publication_detected_in_petdb"
                ],
                "exact_name_collision_n": int(rgr["petdb_exact_name_collision"].sum()),
                "shared_feature_max": int(rgr["observed_feature_n"].max()),
                "status": "strict_external_candidate",
            },
        ]
    )
    summary.to_csv(OUT / "arc_cib_external_candidate_audit.csv", index=False)
    manifest = {
        "feature_contract": FEATURES,
        "publication_overlap_checks": publication_checks,
        "pangaea_922011": pangaea_audit,
        "rio_grande_rift": rgr_audit,
        "deduplication_policy": (
            "Require publication-title/DOI independence from PetDB. Exact normalized sample "
            "names are also checked; short names such as E15 are treated as ambiguous collisions "
            "when PetDB coordinates and citations identify a different province."
        ),
        "label_policy": (
            "Erromango and Vulcan Seamount are ARC-front tests. Vate Trough is retained as "
            "BACK_ARC_TRANSITION and is never scored as ARC or MORB. Rio Grande Rift and "
            "Jemez Lineament basalt samples are CIB, with the project's SiO2=40-55 wt% gate."
        ),
    }
    (OUT / "arc_cib_external_candidate_audit.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
